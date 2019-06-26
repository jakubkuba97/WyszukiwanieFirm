import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side


def set_first(nazwa_pliku):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['B1'] = ""
    sheet.append(["", "NAZWA FIRMY", "ADRES FIRMY", "NUMER TELEFONU", "NUMER NIP", "REPREZENTACJA", "OPERATOR", "ILE SIM", "DATA UMOWY", "UWAGI"])
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 10
    sheet.column_dimensions['I'].width = 15
    sheet.column_dimensions['J'].width = 20
    for row in sheet.iter_rows():
        for cell in row:
            if 'A' not in str(cell.coordinate):
                if str(cell.coordinate)[-1:] != '1':
                    sheet[str(cell.coordinate)].fill = PatternFill(fgColor="000000FF", fill_type="solid")
                sheet[str(cell.coordinate)].font = Font(color='00FFFF00', bold=True)

    workbook.save(nazwa_pliku + ".xlsx")
    workbook.close()


def write_to_file(nazwa_pliku, objects_array, max_length, site_size):
    workbook = openpyxl.load_workbook(nazwa_pliku + ".xlsx")
    sheet = workbook.active

    count = -20
    times_printed = 1
    print("\tZapisano 1 strone! " + str(times_printed) + " / " + str(max_length) + ".")
    for obj in objects_array:
        sheet.append(["", obj.nazwa_firmy, obj.adres_firmy, obj.numer_telefonu, obj.numer_nip, obj.osoby_reprezentujace, obj.operator, obj.ile_sim, obj.data_umowy, obj.uwagi])
        count += 1
        if count > 0 and count % site_size == 0:
            times_printed += 1
            print("\tZapisano 1 strone! " + str(times_printed) + " / " + str(max_length) + ".")

    workbook.save(nazwa_pliku + ".xlsx")
    workbook.close()


def adjust(nazwa_pliku):
    workbook = openpyxl.load_workbook(nazwa_pliku + ".xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
            cell.alignment = cell.alignment.copy(wrapText=True)
            if 'E' in cell.coordinate and cell.coordinate != 'E2':
                sheet[str(cell.coordinate)].font = Font(italic=True)
            if int(str(cell.coordinate)[1:]) % 10 == 0 and 'A' not in str(cell.coordinate):
                sheet[str(cell.coordinate)].fill = PatternFill(fgColor="0000FF00", fill_type="solid")
            if 'B' in cell.coordinate and cell.coordinate != 'B2':
                sheet[str(cell.coordinate)].font = Font(bold=True)
            if 'B' in cell.coordinate and cell.coordinate != 'B1':                              # left
                sheet[str(cell.coordinate)].border = Border(left=Side(style='thick'))
            elif 'J' in cell.coordinate and cell.coordinate != 'J1':                            # right
                sheet[str(cell.coordinate)].border = Border(right=Side(style='thick'))
            if cell.coordinate[1:] == '2' and cell.coordinate[:1] != 'A':                       # up
                if 'B' in cell.coordinate and cell.coordinate != 'B1':
                    sheet[str(cell.coordinate)].border = Border(left=Side(style='thick'), top=Side(style='thick'))
                elif 'J' in cell.coordinate and cell.coordinate != 'J1':
                    sheet[str(cell.coordinate)].border = Border(right=Side(style='thick'), top=Side(style='thick'))
                else:                                                                           # not a corner
                    sheet[str(cell.coordinate)].border = Border(top=Side(style='thick'))
            elif cell.coordinate[1:] == str(sheet.max_row) and cell.coordinate[:1] != 'A':      # down
                if 'B' in cell.coordinate and cell.coordinate != 'B1':
                    sheet[str(cell.coordinate)].border = Border(left=Side(style='thick'), bottom=Side(style='thick'))
                elif 'J' in cell.coordinate and cell.coordinate != 'J1':
                    sheet[str(cell.coordinate)].border = Border(right=Side(style='thick'), bottom=Side(style='thick'))
                else:                                                                           # not a corner
                    sheet[str(cell.coordinate)].border = Border(bottom=Side(style='thick'))
    workbook.save(nazwa_pliku + ".xlsx")
    workbook.close()
